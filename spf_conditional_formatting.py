"""
FI1BBSF05 - Spreadsheet Fundamentals
Lesson 1.4 – Conditional Formatting: Activity 2 Solution
=========================================================
This script loads the provided student scores dataset and applies all
10 conditional formatting requirements using openpyxl, then saves the
result as a new workbook so the original file is preserved.

Run:  python spf_conditional_formatting.py
Output file: SPF_Activity2_CF_Solution.xlsx
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    IconSetRule,
    FormulaRule,
)

# ──────────────────────────────────────────────────────────────
# 1. Load workbook and pick the active sheet
# ──────────────────────────────────────────────────────────────
SOURCE = "SPF+-+Sample+dataset+Activity+2.xlsx"
OUTPUT = "SPF_Activity2_CF_Solution.xlsx"

wb = load_workbook(SOURCE)
ws = wb.active
print(f"Loaded sheet: '{ws.title}'  ({ws.max_row - 1} student rows)")

# ──────────────────────────────────────────────────────────────
# 2. Inject a few duplicate names so rule 8 is visible
#    (rows are 1-indexed; row 1 is the header)
# ──────────────────────────────────────────────────────────────
ws["A3"] = "Student 1"   # Originally "Student 2" → now a duplicate
ws["A10"] = "Student 5"  # Originally "Student 9" → now a duplicate
print("Added duplicate names: 'Student 1' (A3) and 'Student 5' (A10)")

# ──────────────────────────────────────────────────────────────
# 3. Helper fills and fonts
# ──────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def font(hex_color, bold=True):
    return Font(color=hex_color, bold=bold)

# ──────────────────────────────────────────────────────────────
# 4. Range addresses (header is row 1; data rows 2-31)
# ──────────────────────────────────────────────────────────────
last_row = ws.max_row          # 31 (header + 30 students)
name_range = f"A2:A{last_row}"
math_range = f"B2:B{last_row}"
sci_range  = f"C2:C{last_row}"
eng_range  = f"D2:D{last_row}"

print(f"Applying rules to rows 2–{last_row}  (columns A-D)")

# ──────────────────────────────────────────────────────────────
# TASK 1 – Maths < 60  →  solid red background
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    math_range,
    CellIsRule(operator="lessThan", formula=["60"], fill=fill("FF0000")),
)
print("Rule 1 applied – Maths < 60 → red")

# ──────────────────────────────────────────────────────────────
# TASK 2 – Science > 90  →  solid green background
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    sci_range,
    CellIsRule(operator="greaterThan", formula=["90"], fill=fill("00B050")),
)
print("Rule 2 applied – Science > 90 → green")

# ──────────────────────────────────────────────────────────────
# TASK 3 – English between 70 and 80 (inclusive)  →  yellow
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    eng_range,
    CellIsRule(operator="between", formula=["70", "80"], fill=fill("FFFF00")),
)
print("Rule 3 applied – English 70–80 → yellow")

# ──────────────────────────────────────────────────────────────
# TASK 4 – Maths: 3-colour scale  (red → yellow → green)
#          Lower scores = more red; higher scores = more green
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    math_range,
    ColorScaleRule(
        start_type="min",  start_color="FF0000",   # red   (lowest)
        mid_type="percentile", mid_value=50, mid_color="FFFF00",  # yellow (midpoint)
        end_type="max",    end_color="00FF00",      # green (highest)
    ),
)
print("Rule 4 applied – Maths 3-colour scale (red → yellow → green)")

# ──────────────────────────────────────────────────────────────
# TASK 5 – Science: 2-colour scale  (light-pink → dark-blue)
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    sci_range,
    ColorScaleRule(
        start_type="min", start_color="FFAAAA",  # pale red (lowest)
        end_type="max",   end_color="0000FF",    # blue     (highest)
    ),
)
print("Rule 5 applied – Science 2-colour scale (pink → blue)")

# ──────────────────────────────────────────────────────────────
# TASK 6 – English: data bars (blue)
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    eng_range,
    DataBarRule(
        start_type="min", start_value=0,
        end_type="max",   end_value=100,
        color="4472C4",   # standard Excel blue
    ),
)
print("Rule 6 applied – English data bars (blue)")

# ──────────────────────────────────────────────────────────────
# TASK 7 – Maths: 3-arrow icon set
#          ↑ green  = top third   (≥ 67th percentile)
#          → yellow = middle third
#          ↓ red    = bottom third (< 33rd percentile)
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    math_range,
    IconSetRule(
        icon_style="3Arrows",
        type="percent",
        values=[0, 33, 67],   # percentage thresholds
        showValue=True,
    ),
)
print("Rule 7 applied – Maths 3-arrow icon set")

# ──────────────────────────────────────────────────────────────
# TASK 8 – Duplicate names in Name column  →  orange/pink highlight
#          Uses COUNTIF: if the name appears > 1 time → colour it
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    name_range,
    FormulaRule(
        formula=[f"COUNTIF($A$2:$A${last_row},A2)>1"],
        fill=fill("FFC7CE"),        # light red (Excel's default for duplicates)
        font=font("9C0006"),        # dark red text
    ),
)
print("Rule 8 applied – Duplicate names → pink/dark-red")

# ──────────────────────────────────────────────────────────────
# TASK 9 – Top 3 Science scores  →  green background, bold text
#          LARGE(range, 3) returns the 3rd highest value
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    sci_range,
    FormulaRule(
        formula=[f"C2>=LARGE($C$2:$C${last_row},3)"],
        fill=fill("C6EFCE"),        # pale green
        font=font("276221"),        # dark green text
    ),
)
print("Rule 9 applied – Top 3 Science scores → green")

# ──────────────────────────────────────────────────────────────
# TASK 10 – Bottom 3 Maths scores  →  red background, bold text
#           SMALL(range, 3) returns the 3rd lowest value
# ──────────────────────────────────────────────────────────────
ws.conditional_formatting.add(
    math_range,
    FormulaRule(
        formula=[f"B2<=SMALL($B$2:$B${last_row},3)"],
        fill=fill("FFD7CE"),        # pale red/orange
        font=font("9C0006"),        # dark red text
    ),
)
print("Rule 10 applied – Bottom 3 Maths scores → red/orange")

# ──────────────────────────────────────────────────────────────
# 5. Style the header row for clarity
# ──────────────────────────────────────────────────────────────
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# ──────────────────────────────────────────────────────────────
# 6. Auto-fit column widths (approximate)
# ──────────────────────────────────────────────────────────────
col_widths = {"A": 14, "B": 14, "C": 12, "D": 12}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# ──────────────────────────────────────────────────────────────
# 7. Add a legend sheet so you can see what each rule does
# ──────────────────────────────────────────────────────────────
legend = wb.create_sheet("Legend")
rules = [
    ("Task",  "Column",      "Rule description",                        "Formatting"),
    ("1",     "Mathematics", "Score < 60",                              "Solid red background"),
    ("2",     "Science",     "Score > 90",                              "Solid green background"),
    ("3",     "English",     "Score between 70 and 80 (incl.)",         "Solid yellow background"),
    ("4",     "Mathematics", "Colour scale – low → high",               "Red → Yellow → Green градиент"),
    ("5",     "Science",     "2-colour scale – low → high",             "Pink → Blue gradient"),
    ("6",     "English",     "Data bars",                               "Blue bar proportional to value"),
    ("7",     "Mathematics", "Icon set (3 Arrows)",                     "↑ top third  →  mid  ↓ bottom"),
    ("8",     "Name",        "Duplicate entry",                         "Pink background, dark-red text"),
    ("9",     "Science",     "Top 3 scores  (≥ 3rd highest value)",     "Green background, bold text"),
    ("10",    "Mathematics", "Bottom 3 scores  (≤ 3rd lowest value)",   "Red/orange background, bold text"),
    ("11",    "All",         "Dynamic – change any score and save",      "Rules recalculate automatically"),
]
for row_data in rules:
    legend.append(list(row_data))

# Style legend header
for cell in legend[1]:
    cell.fill = header_fill
    cell.font = header_font

for col, width in {"A": 6, "B": 14, "C": 44, "D": 38}.items():
    legend.column_dimensions[col].width = width

# ──────────────────────────────────────────────────────────────
# 8. Save
# ──────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"\n✓  Saved: {OUTPUT}")
print("  Open it in Excel or LibreOffice Calc to see all formatting.")
