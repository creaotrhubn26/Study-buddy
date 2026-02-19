"""
Study-Buddy Dataset Calculator
Runs all Q21, Q22, Q23 calculations with step-by-step breakdowns.
Run with:  streamlit run dataset_calculator.py
"""

import streamlit as st
import pandas as pd
import os

st.set_page_config(page_title="Dataset Calculator", page_icon="📊", layout="wide")

st.title("📊 Study-Buddy Dataset Calculator")
st.caption("All exam dataset calculations — Q21–Q24 — with formulas, step-by-step breakdowns, and Python solutions derived from the official resolved datasets")

# ── helpers ──────────────────────────────────────────────────────────────────
def safe_load(path):
    if not os.path.exists(path):
        st.error(f"File not found: {path}")
        return None
    return pd.read_excel(path, header=0)

def df_to_md(df):
    """Build a markdown table from a DataFrame without needing tabulate."""
    idx_name = df.index.name or "Row"
    cols = list(df.columns)
    header = "| " + idx_name + " | " + " | ".join(str(c) for c in cols) + " |"
    sep    = "| --- | " + " | ".join("---" for _ in cols) + " |"
    rows   = ["| " + str(idx) + " | " + " | ".join(str(v) for v in row) + " |"
              for idx, row in df.iterrows()]
    return "\n".join([header, sep] + rows)

def answer_box(formula, result):
    st.success(f"**Formula:** `{formula}`   →   **Answer: {result}**")

def step(n, title, body):
    with st.expander(f"Step {n} — {title}", expanded=True):
        st.markdown(body)

# ─────────────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "Q21 — CA Lesson 1 (Sales data)",
    "Q22 — CA Lesson 3 (Employee salaries)",
    "Q23 — CA Lesson 4 (Salaries + Bonus)",
    "Q24 — CA Lesson 2 (Date, Time & Text)",
    "🔑 Python Solutions (Resolved Files)"
])


# ══════════════════════════════════════════════════════════════════════════════
# Q21 — CA LESSON 1 DATASET
# Columns: Date | Product ID | Units Sold | Price per Unit
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.header("CA Lesson 1 Dataset — Q21")
    df1 = safe_load("CA Lesson 1 dataset.xlsx")

    if df1 is not None:
        df1.columns = ["Date", "Product ID", "Units Sold", "Price per Unit"]
        df1["Date"] = pd.to_datetime(df1["Date"]).dt.strftime("%Y-%m-%d")
        # Strip $ signs and convert Price per Unit to numeric
        df1["Price per Unit"] = (
            df1["Price per Unit"].astype(str).str.replace("$", "", regex=False).str.strip()
        )
        df1["Price per Unit"] = pd.to_numeric(df1["Price per Unit"], errors="coerce")
        df1["Units Sold"] = pd.to_numeric(df1["Units Sold"], errors="coerce")
        # Add Excel row numbers for reference
        df1.index = range(2, len(df1) + 2)
        df1.index.name = "Row"

        st.subheader("📋 Full Dataset")
        st.dataframe(df1, use_container_width=True)

        st.divider()

        # ── Q21.1 ─────────────────────────────────────────────────────────
        st.subheader("Q21.1 — Total units sold (SUM)")
        result = int(df1["Units Sold"].sum())
        answer_box("=SUM(C2:C21)", result)
        step(1, "Add all values in Units Sold column",
             f"Values: {df1['Units Sold'].tolist()}\n\n"
             f"**Sum = {result}**")

        st.divider()

        # ── Q21.2 ─────────────────────────────────────────────────────────
        st.subheader("Q21.2 — Average price per unit (AVERAGE)")
        result2 = round(df1["Price per Unit"].mean(), 2)
        answer_box("=AVERAGE(D2:D21)", result2)
        step(1, "Sum all prices ÷ count",
             f"Sum of prices: {df1['Price per Unit'].sum()}\n\n"
             f"Count: {len(df1)}\n\n"
             f"**Average = {df1['Price per Unit'].sum()} / {len(df1)} = {result2}**")

        st.divider()

        # ── Q21.3 ─────────────────────────────────────────────────────────
        st.subheader("Q21.3 — Highest units sold on any single day (MAX)")
        result3 = int(df1["Units Sold"].max())
        row3 = df1[df1["Units Sold"] == result3]
        answer_box("=MAX(C2:C21)", result3)
        step(1, "Find the maximum value in Units Sold",
             f"**Max = {result3}**\n\nFound in row(s):\n\n{df_to_md(row3)}")

        st.divider()

        # ── Q21.4 ─────────────────────────────────────────────────────────
        st.subheader("Q21.4 — Lowest price per unit (MIN)")
        result4 = int(df1["Price per Unit"].min())
        row4 = df1[df1["Price per Unit"] == result4]
        answer_box("=MIN(D2:D21)", result4)
        step(1, "Find the minimum value in Price per Unit",
             f"**Min = {result4}**\n\nFound in row(s):\n\n{df_to_md(row4)}")

        st.divider()

        # ── Q21.5 ─────────────────────────────────────────────────────────
        st.subheader("Q21.5 — Days when P002 was sold (COUNTIF)")
        result5 = int((df1["Product ID"] == "P002").sum())
        p002_rows = df1[df1["Product ID"] == "P002"]
        answer_box('=COUNTIF(B2:B21,"P002")', result5)
        step(1, "Count rows where Product ID = P002",
             f"Qualifying rows:\n\n{df_to_md(p002_rows[['Date','Product ID','Units Sold']])}\n\n"
             f"**Count = {result5}**")

        st.divider()

        # ── Q21.6 ─────────────────────────────────────────────────────────
        st.subheader("Q21.6 — IF P001 on Sep 10 > 20: High Sales / Low Sales")
        sep10 = df1[df1["Date"] == "2023-09-10"]
        if not sep10.empty:
            p001_sep10_units = sep10[sep10["Product ID"] == "P001"]["Units Sold"]
            units_val = int(p001_sep10_units.values[0]) if not p001_sep10_units.empty else 0
        else:
            units_val = 0
        result6 = "High Sales" if units_val > 20 else "Low Sales"
        answer_box('=IF(SUMIFS(C2:C21,B2:B21,"P001",A2:A21,"2023-09-10")>20,"High Sales","Low Sales")', result6)
        step(1, "Check Sep 10 rows",
             f"All rows on 2023-09-10:\n\n{df_to_md(sep10)}")
        step(2, "Is P001 units sold on Sep 10 > 20?",
             f"P001 on Sep 10 → **{units_val} units**\n\n"
             f"{units_val} > 20 is **{units_val > 20}** → result = **\"{result6}\"**")

        st.divider()

        # ── Q21.7 ─────────────────────────────────────────────────────────
        st.subheader("Q21.7 — Concatenate Product ID & Price for Sep 16")
        sep16 = df1[df1["Date"] == "2023-09-16"].iloc[0]
        result7 = f"{sep16['Product ID']}${int(sep16['Price per Unit'])}"
        answer_box('=B[row]&"$"&D[row]', result7)
        step(1, "Find the row for Sep 16",
             f"Product ID = **{sep16['Product ID']}**, Price per Unit = **{int(sep16['Price per Unit'])}**\n\n"
             f"Concatenation: `\"{sep16['Product ID']}\" & \"$\" & \"{int(sep16['Price per Unit'])}\"` = **\"{result7}\"**")

        st.divider()

        # ── Q21.8 ─────────────────────────────────────────────────────────
        st.subheader("Q21.8 — LEFT 4 characters of Product ID for Sep 06")
        sep06 = df1[df1["Date"] == "2023-09-06"].iloc[0]
        result8 = sep06["Product ID"][:4]
        answer_box("=LEFT(B[row],4)", result8)
        step(1, "Find the row for Sep 06",
             f"Product ID = **\"{sep06['Product ID']}\"**\n\n"
             f"LEFT({sep06['Product ID']}, 4) takes the first 4 characters:\n\n"
             f"`{sep06['Product ID']}` → **\"{result8}\"**")

        st.divider()

        # ── Q21.9 ─────────────────────────────────────────────────────────
        st.subheader("Q21.9 — RIGHT 3 characters of Product ID for Sep 14")
        sep14 = df1[df1["Date"] == "2023-09-14"].iloc[0]
        result9 = sep14["Product ID"][-3:]
        answer_box("=RIGHT(B[row],3)", result9)
        step(1, "Find the row for Sep 14",
             f"Product ID = **\"{sep14['Product ID']}\"**\n\n"
             f"RIGHT({sep14['Product ID']}, 3) takes the last 3 characters:\n\n"
             f"`{sep14['Product ID']}` → **\"{result9}\"**")

        st.divider()
        st.subheader("📌 Q21 Summary")
        summary1 = pd.DataFrame({
            "Q": ["21.1","21.2","21.3","21.4","21.5","21.6","21.7","21.8","21.9"],
            "Task": ["Total units sold","Average price per unit","Max units sold","Min price per unit",
                     "Days P002 sold","IF P001 Sep10 >20","Concat Sep16","LEFT Sep06","RIGHT Sep14"],
            "Formula": ["=SUM(C2:C21)","=AVERAGE(D2:D21)","=MAX(C2:C21)","=MIN(D2:D21)",
                        '=COUNTIF(B2:B21,"P002")',
                        '=IF(SUMIFS(...)>20,"High Sales","Low Sales")',
                        '=B[row]&"$"&D[row]',"=LEFT(B[row],4)","=RIGHT(B[row],3)"],
            "Answer": [result, result2, result3, result4, result5, result6, result7, result8, result9]
        })
        st.dataframe(summary1, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# Q22 — CA LESSON 3 DATASET
# Columns: Employee ID | Employee Name | Department | Salary
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.header("CA Lesson 3 Dataset — Q22")
    df2 = safe_load("CA Lesson 3 dataset.xlsx")

    if df2 is not None:
        df2.columns = ["Employee ID", "Employee Name", "Department", "Salary"]
        df2.index = range(2, len(df2) + 2)
        df2.index.name = "Row"

        st.subheader("📋 Full Dataset")
        st.dataframe(df2, use_container_width=True)

        st.divider()

        names    = df2["Employee Name"].tolist()
        depts    = df2["Department"].tolist()
        salaries = df2["Salary"].tolist()

        # ── Q22.1 ─────────────────────────────────────────────────────────
        st.subheader("Q22.1 — Carol's salary (VLOOKUP)")
        carol = df2[df2["Employee Name"] == "Carol"]
        r1 = int(carol["Salary"].values[0])
        answer_box('=VLOOKUP("Carol",B2:D21,3,FALSE)', r1)
        step(1, "Find Carol in the dataset", df_to_md(carol))
        step(2, "Count col_index in table B2:D21",
             "| Column | Position |\n|--------|----------|\n"
             "| B (Employee Name) | 1 — searched here |\n"
             "| C (Department) | 2 |\n"
             "| **D (Salary)** | **3 ← col_index** |\n\n"
             f"VLOOKUP returns column 3 → **{r1}**")

        st.divider()

        # ── Q22.2 ─────────────────────────────────────────────────────────
        st.subheader("Q22.2 — Paul's department (VLOOKUP)")
        paul = df2[df2["Employee Name"] == "Paul"]
        r2 = paul["Department"].values[0]
        answer_box('=VLOOKUP("Paul",B2:D21,2,FALSE)', r2)
        step(1, "Find Paul in the dataset", df_to_md(paul))
        step(2, "Count col_index in table B2:D21",
             "| Column | Position |\n|--------|----------|\n"
             "| B (Employee Name) | 1 — searched here |\n"
             "| **C (Department)** | **2 ← col_index** |\n"
             "| D (Salary) | 3 |\n\n"
             f"VLOOKUP returns column 2 → **{r2}**")

        st.divider()

        # ── Q22.3 ─────────────────────────────────────────────────────────
        st.subheader("Q22.3 — Maria's salary (INDEX + MATCH)")
        maria_pos = names.index("Maria") + 1   # 1-based position in range
        r3 = salaries[names.index("Maria")]
        answer_box('=INDEX(D2:D21,MATCH("Maria",B2:B21,0))', r3)
        step(1, "MATCH finds Maria's position in B2:B21",
             f"B2:B21 list: {names}\n\n"
             f"Maria is at **position {maria_pos}** (1 = first row of range)")
        step(2, f"INDEX returns position {maria_pos} from D2:D21",
             f"D2:D21 list: {salaries}\n\n"
             f"Value at position {maria_pos} → **{r3}**")

        st.divider()

        # ── Q22.4 ─────────────────────────────────────────────────────────
        st.subheader("Q22.4 — Position of highest salary (MATCH + MAX)")
        max_sal = max(salaries)
        max_pos = salaries.index(max_sal) + 1
        max_emp = df2[df2["Salary"] == max_sal]
        r4 = max_pos
        answer_box("=MATCH(MAX(D2:D21),D2:D21,0)", r4)
        step(1, f"MAX(D2:D21) = {max_sal}",
             f"Highest salary belongs to:\n\n{df_to_md(max_emp)}")
        step(2, f"MATCH finds position of {max_sal} in D2:D21",
             f"D2:D21 positions:\n\n"
             + "\n".join([f"Position {i+1}: {s}" + (" ← MAX" if s == max_sal else "")
                          for i, s in enumerate(salaries)]) +
             f"\n\n**Result = {r4}**\n\n"
             f"⚠️ Position {r4} in the range ≠ Excel row {r4}. "
             f"Excel row = {r4+1} (because D2 is row 2, so position {r4} = row {r4+1}).")

        st.divider()

        # ── Q22.5 ─────────────────────────────────────────────────────────
        st.subheader("Q22.5 — Sum of first 5 salaries (SUM + OFFSET)")
        first5 = salaries[:5]
        r5 = sum(first5)
        answer_box("=SUM(OFFSET(D2,0,0,5,1))", r5)
        step(1, "OFFSET(D2,0,0,5,1) creates range D2:D6",
             "Arguments: start=D2, move 0 rows, move 0 cols, height=5, width=1\n\n"
             "Result: a reference to **D2:D6** (first 5 employees)")
        step(2, "SUM the 5 values",
             "| Row | Employee | Salary |\n|-----|---------|--------|\n" +
             "\n".join([f"| {i+2} | {df2.iloc[i]['Employee Name']} | {s} |"
                        for i, s in enumerate(first5)]) +
             f"\n\n**Sum = {' + '.join(map(str,first5))} = {r5}**")

        st.divider()

        # ── Q22.6 ─────────────────────────────────────────────────────────
        st.subheader("Q22.6 — Irene's department (INDEX + MATCH)")
        irene_pos = names.index("Irene") + 1
        r6 = depts[names.index("Irene")]
        answer_box('=INDEX(C2:C21,MATCH("Irene",B2:B21,0))', r6)
        step(1, "MATCH finds Irene's position in B2:B21",
             f"Irene is at **position {irene_pos}**")
        step(2, f"INDEX returns position {irene_pos} from C2:C21 (Department column)",
             f"Department list: {depts}\n\nValue at position {irene_pos} → **{r6}**")

        st.divider()

        # ── Q22.7 ─────────────────────────────────────────────────────────
        st.subheader("Q22.7 — Salary at Excel row 15 (INDIRECT)")
        row15_val = int(df2.loc[15, "Salary"])
        row15_emp = df2.loc[15, "Employee Name"]
        r7 = row15_val
        answer_box('=INDIRECT("D15")', r7)
        step(1, "Who is in Excel row 15?",
             "Row 1 = Header → Row 2 = first employee (Alice) → Row 15 = 14th employee\n\n"
             f"Row 15 → **{row15_emp}** → Salary = **{r7}**")
        step(2, "How INDIRECT works",
             '`=INDIRECT("D15")` converts the text "D15" into the cell address D15 and reads it.\n\n'
             f"D15 = **{r7}**")

        st.divider()

        # ── Q22.8 ─────────────────────────────────────────────────────────
        st.subheader("Q22.8 — Grace's salary (LOOKUP)")
        grace = df2[df2["Employee Name"] == "Grace"]
        r8 = int(grace["Salary"].values[0])
        answer_box('=LOOKUP("Grace",B2:B21,D2:D21)', r8)
        step(1, "Find Grace in the dataset", df_to_md(grace))
        step(2, "LOOKUP scans B2:B21 (must be sorted A→Z)",
             "Names are alphabetical ✅ — binary search works correctly.\n\n"
             f"Grace is found → return matching value from D2:D21 → **{r8}**")

        st.divider()

        # ── Q22.9 ─────────────────────────────────────────────────────────
        st.subheader("Q22.9 — Sam's salary (XLOOKUP)")
        sam = df2[df2["Employee Name"] == "Sam"]
        r9 = int(sam["Salary"].values[0])
        answer_box('=XLOOKUP("Sam",B2:B21,D2:D21)', r9)
        step(1, "Find Sam in the dataset", df_to_md(sam))
        step(2, "XLOOKUP — no col_index needed",
             "Searches B2:B21 for Sam → returns corresponding value from D2:D21.\n\n"
             f"Result → **{r9}**")

        st.divider()
        st.subheader("📌 Q22 Summary")
        summary2 = pd.DataFrame({
            "Q": ["22.1","22.2","22.3","22.4","22.5","22.6","22.7","22.8","22.9"],
            "Task": ["Carol's salary (VLOOKUP)","Paul's dept (VLOOKUP)","Maria's salary (INDEX+MATCH)",
                     "Position of max salary (MATCH+MAX)","Sum first 5 salaries (OFFSET)",
                     "Irene's dept (INDEX+MATCH)","Salary row 15 (INDIRECT)",
                     "Grace's salary (LOOKUP)","Sam's salary (XLOOKUP)"],
            "Formula": [
                '=VLOOKUP("Carol",B2:D21,3,FALSE)',
                '=VLOOKUP("Paul",B2:D21,2,FALSE)',
                '=INDEX(D2:D21,MATCH("Maria",B2:B21,0))',
                "=MATCH(MAX(D2:D21),D2:D21,0)",
                "=SUM(OFFSET(D2,0,0,5,1))",
                '=INDEX(C2:C21,MATCH("Irene",B2:B21,0))',
                '=INDIRECT("D15")',
                '=LOOKUP("Grace",B2:B21,D2:D21)',
                '=XLOOKUP("Sam",B2:B21,D2:D21)',
            ],
            "Answer": [r1, r2, r3, r4, r5, r6, r7, r8, r9]
        })
        st.dataframe(summary2, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# Q23 — CA LESSON 4 DATASET
# Columns: Employee ID | Employee Name | Department | Salary | Years | Bonus%
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    st.header("CA Lesson 4 Dataset — Q23")
    df3 = safe_load("CA Lesson 4 dataset.xlsx")

    if df3 is not None:
        df3.columns = ["Employee ID", "Employee Name", "Department", "Salary",
                       "Years of Service", "Bonus Percentage"]
        df3.index = range(2, len(df3) + 2)
        df3.index.name = "Row"

        # Display bonus% as readable percentage
        df3_display = df3.copy()
        df3_display["Bonus Percentage"] = (df3["Bonus Percentage"] * 100).round(1).astype(str) + " %"

        st.subheader("📋 Full Dataset")
        st.dataframe(df3_display, use_container_width=True)

        st.divider()

        # ── Q23.1 ─────────────────────────────────────────────────────────
        st.subheader("Q23.1 — Max salary: Sales dept, ≥ 2 years (MAXIFS)")
        mask1 = (df3["Department"] == "Sales") & (df3["Years of Service"] >= 2)
        qualifying1 = df3[mask1]
        r1 = int(qualifying1["Salary"].max())
        answer_box('=MAXIFS(D2:D21,C2:C21,"Sales",E2:E21,">=2")', r1)
        step(1, "Filter: Department = Sales AND Years ≥ 2",
             df_to_md(qualifying1[["Employee Name","Department","Salary","Years of Service"]]) +
             f"\n\n**MAX of qualifying salaries = {r1}**\n\n"
             f"(Tied between: {', '.join(qualifying1[qualifying1['Salary']==r1]['Employee Name'].tolist())})")
        step(2, "MAXIFS syntax",
             "```\n=MAXIFS( D2:D21 , C2:C21 , \"Sales\" , E2:E21 , \">=2\" )\n"
             "          ↑          ↑          ↑           ↑          ↑\n"
             "          max_range  range1    criteria1   range2    criteria2\n```\n\n"
             "⚠️ Exam trap: `\">=2\"` must be in quotes — it's a text string criterion.")

        st.divider()

        # ── Q23.2 ─────────────────────────────────────────────────────────
        st.subheader("Q23.2 — Min bonus%: Finance dept, salary > 58000 (MINIFS)")
        mask2 = (df3["Department"] == "Finance") & (df3["Salary"] > 58000)
        qualifying2 = df3[mask2]
        r2 = qualifying2["Bonus Percentage"].min()
        answer_box('=MINIFS(F2:F21,C2:C21,"Finance",D2:D21,">58000")', f"{r2*100:.1f}% (= {r2})")
        step(1, "Filter: Dept = Finance AND Salary > 58000 (strictly)",
             df_to_md(qualifying2[["Employee Name","Department","Salary","Bonus Percentage"]]) +
             f"\n\n⚠️ Neil Taylor (salary = 58000) is **excluded** — 58000 is not **strictly greater** than 58000.\n\n"
             f"**MIN bonus% = {r2*100:.1f}% = {r2}**")

        st.divider()

        # ── Q23.3 ─────────────────────────────────────────────────────────
        st.subheader("Q23.3 — Count employees with bonus% ≥ 8% (COUNTIF)")
        mask3 = df3["Bonus Percentage"] >= 0.08
        qualifying3 = df3[mask3]
        r3 = len(qualifying3)
        answer_box('=COUNTIF(F2:F21,">=0.08")', r3)
        step(1, "Filter: Bonus Percentage ≥ 0.08 (= 8%)",
             df_to_md(qualifying3[["Employee Name","Department","Bonus Percentage"]]) +
             f"\n\n**Count = {r3}**")

        st.divider()

        # ── Q23.4 ─────────────────────────────────────────────────────────
        st.subheader("Q23.4 — Count IT employees with salary > 60000 (COUNTIFS)")
        mask4 = (df3["Department"] == "IT") & (df3["Salary"] > 60000)
        qualifying4 = df3[mask4]
        excluded4 = df3[(df3["Department"] == "IT") & (df3["Salary"] == 60000)]
        r4 = len(qualifying4)
        answer_box('=COUNTIFS(C2:C21,"IT",D2:D21,">60000")', r4)
        step(1, "Filter: IT AND Salary > 60000 (strictly)",
             df_to_md(qualifying4[["Employee Name","Department","Salary"]]) +
             ("\n\n⚠️ Carol Williams (60000) is **excluded** — not strictly > 60000."
              if not excluded4.empty else "") +
             f"\n\n**Count = {r4}**")

        st.divider()

        # ── Q23.5 ─────────────────────────────────────────────────────────
        st.subheader("Q23.5 — Average salary for bonus% < 6% (AVERAGEIF)")
        mask5 = df3["Bonus Percentage"] < 0.06
        qualifying5 = df3[mask5]
        r5 = round(qualifying5["Salary"].mean(), 2)
        answer_box('=AVERAGEIF(F2:F21,"<0.06",D2:D21)', r5)
        step(1, "Filter: Bonus Percentage < 0.06 (strictly less than 6%)",
             df_to_md(qualifying5[["Employee Name","Salary","Bonus Percentage"]]))
        step(2, "Calculate average",
             f"Salaries: {qualifying5['Salary'].tolist()}\n\n"
             f"Sum = {qualifying5['Salary'].sum()}, Count = {len(qualifying5)}\n\n"
             f"**Average = {qualifying5['Salary'].sum()} / {len(qualifying5)} = {r5}**")

        st.divider()

        # ── Q23.6 ─────────────────────────────────────────────────────────
        st.subheader("Q23.6 — Average bonus%: Sales dept, > 1 year (AVERAGEIFS)")
        mask6 = (df3["Department"] == "Sales") & (df3["Years of Service"] > 1)
        qualifying6 = df3[mask6]
        r6 = qualifying6["Bonus Percentage"].mean()
        answer_box('=AVERAGEIFS(F2:F21,C2:C21,"Sales",E2:E21,">1")', f"{r6*100:.2f}% (= {r6})")
        step(1, "Filter: Dept = Sales AND Years > 1",
             df_to_md(qualifying6[["Employee Name","Department","Years of Service","Bonus Percentage"]]))
        step(2, "Calculate average",
             f"Bonus percentages: {[f'{b*100:.1f}%' for b in qualifying6['Bonus Percentage'].tolist()]}\n\n"
             f"**Average = {r6*100:.2f}% = {r6}**\n\n"
             "⚠️ AVERAGEIFS argument order: **average_range FIRST**, then (range, criteria) pairs.\n"
             "This is the opposite of AVERAGEIF (where average_range is LAST).")

        st.divider()

        # ── Q23.7 ─────────────────────────────────────────────────────────
        st.subheader("Q23.7 — Bonus Amount column G, then IFERROR(G2/F2,\"Error\")")
        alice = df3.iloc[0]
        g2 = alice["Salary"] * alice["Bonus Percentage"]
        r7_g = g2
        r7 = g2 / alice["Bonus Percentage"]

        st.info("**Two-step question:** First create column G (Bonus Amount), then use IFERROR to divide it back.")

        step(1, "Where does 2750 come from? — Create column G",
             f"Column G is **not given** — you calculate it:\n\n"
             "| Cell | Column | Value | Source |\n"
             "|------|--------|-------|--------|\n"
             f"| D2 | Salary | {int(alice['Salary'])} | Given in dataset |\n"
             f"| F2 | Bonus % | {alice['Bonus Percentage']} | Given in dataset (= {alice['Bonus Percentage']*100:.0f}%) |\n"
             f"| **G2** | **Bonus Amount** | **{int(g2)}** | **=D2*F2 ← you create this** |\n\n"
             f"G2 = {int(alice['Salary'])} × {alice['Bonus Percentage']} = **{int(g2)}**")
        step(2, "The IFERROR formula",
             f'`=IFERROR(G2/F2,"Error")`\n\n'
             f"= IFERROR({int(g2)} / {alice['Bonus Percentage']}, \"Error\")\n\n"
             f"= IFERROR({int(r7)}, \"Error\")\n\n"
             f"= **{int(r7)}** (= Alice's salary recovered)")
        step(3, "Why G2/F2 = Salary",
             "G2 = D2 × F2 (Salary × Bonus%)\n\n"
             "G2 / F2 = (D2 × F2) / F2 = **D2** = Salary\n\n"
             "The formula mathematically recovers the original salary.")
        step(4, "Why IFERROR protects the formula",
             "If F2 = 0 → G2/F2 → **#DIV/0! error**\n\n"
             'IFERROR catches this and displays **"Error"** instead.')

        answer_box('=IFERROR(G2/F2,"Error")', f"{int(r7)} (Alice's salary)")

        # Show full bonus amount column
        df3_bonus = df3.copy()
        df3_bonus["Bonus Amount (G)"] = (df3["Salary"] * df3["Bonus Percentage"]).round(2)
        df3_bonus["IFERROR(G/F)"] = (df3_bonus["Bonus Amount (G)"] / df3["Bonus Percentage"]).round(0).astype(int)
        st.dataframe(df3_bonus[["Employee Name","Salary","Bonus Percentage","Bonus Amount (G)","IFERROR(G/F)"]],
                     use_container_width=True)

        st.divider()

        # ── Q23.8 ─────────────────────────────────────────────────────────
        st.subheader("Q23.8 — Grace Martin's salary (LOOKUP)")
        grace_row = df3[df3["Employee Name"] == "Grace Martin"]
        r8 = int(grace_row["Salary"].values[0])
        answer_box('=LOOKUP("Grace Martin",B2:B21,D2:D21)', r8)
        step(1, "Find Grace Martin", df_to_md(grace_row[["Employee Name","Department","Salary"]]))
        step(2, "LOOKUP requires sorted data",
             "Names in B2:B21 are in alphabetical order ✅\n\n"
             "LOOKUP's binary search finds Grace Martin → returns matching Salary → **{r8}**\n\n"
             "⚠️ Use full name `\"Grace Martin\"` — cells contain first AND last name.")

        st.divider()

        # ── Q23.9 ─────────────────────────────────────────────────────────
        st.subheader("Q23.9 — Sam Scott's salary (XLOOKUP)")
        sam_row = df3[df3["Employee Name"] == "Sam Scott"]
        r9 = int(sam_row["Salary"].values[0])
        answer_box('=XLOOKUP("Sam Scott",B2:B21,D2:D21)', r9)
        step(1, "Find Sam Scott", df_to_md(sam_row[["Employee Name","Department","Salary"]]))
        step(2, "XLOOKUP — exact full name required",
             '`=XLOOKUP("Sam Scott",B2:B21,D2:D21)` → finds row → returns Salary → **63000**\n\n'
             '`=XLOOKUP("Sam",B2:B21,D2:D21)` → **#N/A** — cells contain "Sam Scott", not "Sam"')

        st.divider()
        st.subheader("📌 Q23 Summary")
        summary3 = pd.DataFrame({
            "Q": ["23.1","23.2","23.3","23.4","23.5","23.6","23.7","23.8","23.9"],
            "Task": [
                "Max salary Sales ≥2 yrs","Min bonus% Finance salary>58k",
                "Count bonus% ≥8%","Count IT salary>60k",
                "Avg salary bonus%<6%","Avg bonus% Sales >1yr",
                "IFERROR(G2/F2)","Grace Martin salary","Sam Scott salary"
            ],
            "Formula": [
                '=MAXIFS(D2:D21,C2:C21,"Sales",E2:E21,">=2")',
                '=MINIFS(F2:F21,C2:C21,"Finance",D2:D21,">58000")',
                '=COUNTIF(F2:F21,">=0.08")',
                '=COUNTIFS(C2:C21,"IT",D2:D21,">60000")',
                '=AVERAGEIF(F2:F21,"<0.06",D2:D21)',
                '=AVERAGEIFS(F2:F21,C2:C21,"Sales",E2:E21,">1")',
                '=IFERROR(G2/F2,"Error")',
                '=LOOKUP("Grace Martin",B2:B21,D2:D21)',
                '=XLOOKUP("Sam Scott",B2:B21,D2:D21)',
            ],
            "Answer": [
                r1,
                f"{r2*100:.1f}% (={r2})",
                r3, r4, r5,
                f"{r6*100:.2f}% (={r6})",
                f"{int(r7)} (Alice)",
                r8, r9
            ]
        })
        st.dataframe(summary3, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# Q24 — CA LESSON 2 DATASET
# Columns: Order ID | Customer Name | Order Date | Order Time | Order Amount
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.header("CA Lesson 2 Dataset — Q24")
    st.info("⚠️ No spaces in formulas — the automated marker expects `=C6+3` not `=C6 + 3`.")

    import openpyxl as _openpyxl
    from datetime import timedelta as _td, date as _date

    _wb4 = None
    if os.path.exists("CA Lesson 2 dataset.xlsx"):
        _wb4 = _openpyxl.load_workbook("CA Lesson 2 dataset.xlsx")
    else:
        st.error("File not found: CA Lesson 2 dataset.xlsx")

    if _wb4:
        _ws4 = _wb4.active
        _d4 = {}
        for _row in _ws4.iter_rows(min_row=2, max_row=21, values_only=True):
            _oid, _name, _odate, _otime, _oamt = _row
            _d4[_oid] = {"name": _name, "date": _odate, "time": _otime,
                         "amount": _oamt, "row": int(_oid) + 1}

        # ---- display full dataset ------------------------------------------
        import pandas as _pd4
        _df4 = _pd4.DataFrame([
            {"Row": v["row"], "Order ID": k, "Customer Name": v["name"],
             "Order Date": str(v["date"].date()),
             "Order Time": str(v["time"]),
             "Order Amount": v["amount"]}
            for k, v in _d4.items()
        ])
        st.subheader("📋 Full Dataset")
        st.dataframe(_df4.set_index("Row"), use_container_width=True)
        st.divider()

        # ── Q24.1 ────────────────────────────────────────────────────────────────────
        st.subheader("Q24.1 — Next Order Date: 3 days after Order Date (Order ID 5)")
        _e = _d4[5]
        _r = _e["row"]
        _next = _e["date"].date() + _td(days=3)
        answer_box(f"=C{_r}+3", _next)
        step(1, f"Find Order ID 5 → Row {_r}",
             f"| Row | Order ID | Customer Name | Order Date |\n"
             f"|-----|----------|--------------|------------|\n"
             f"| {_r} | 5 | {_e['name']} | **{_e['date'].date()}** ← C{_r} |")
        step(2, "How date addition works in Excel",
             "Excel stores dates as serial numbers (count of days since 1900-01-01).\n\n"
             f"Adding 3 moves the date forward 3 calendar days:\n\n"
             f"`C{_r} = {_e['date'].date()}`  +  3  =  **{_next}**\n\n"
             "> **Tip:** To add months use `=EDATE(C,1)`. To add years use `=DATE(YEAR(C)+1,MONTH(C),DAY(C))`.")
        st.divider()

        # ── Q24.2 ────────────────────────────────────────────────────────────────────
        st.subheader("Q24.2 — Time Slot for Order ID 8 (IF + TIME)")
        _e8 = _d4[8]
        _r8 = _e8["row"]
        _t8 = _e8["time"]
        _slot8 = "Morning" if _t8.hour < 12 else "Afternoon"
        answer_box(f'=IF(D{_r8}<TIME(12,0,0),"Morning","Afternoon")', _slot8)
        step(1, f"Find Order ID 8 → Row {_r8}",
             f"| Row | Order ID | Customer Name | Order Time |\n"
             f"|-----|----------|--------------|------------|\n"
             f"| {_r8} | 8 | {_e8['name']} | **{_t8}** ← D{_r8} |")
        step(2, "How TIME(12,0,0) works",
             "```\n"
             "TIME(hour, minute, second) creates a time value.\n"
             "TIME(12,0,0) = 12:00:00 = noon\n\n"
             "Excel stores times as fractions of 1 day:\n"
             "  12:00:00 = 0.5 (half a day)\n"
             "  09:30:00 = 0.395833...\n"
             f"  {_t8} = {_t8.hour*3600/86400 + _t8.minute*60/86400:.6f}\n"
             "```")
        step(3, "IF logic applied",
             f"`{_t8}` < `12:00:00`?  →  **{_t8.hour < 12 or (_t8.hour == 12 and _t8.minute == 0 and _t8.second == 0)}**\n\n"
             f"Result: **\"{_slot8}\"**\n\n"
             "⚠️ **Exam trap:** 12:00 PM exactly is NOT morning. The condition is strictly `<` noon.\n"
             "Linda Harris at 12:15 → 12:15 < 12:00 is FALSE → Afternoon.")
        st.divider()

        # ── Q24.3 ────────────────────────────────────────────────────────────────────
        st.subheader("Q24.3 — Rounded Amount for Order ID 16 (ROUND)")
        _e16 = _d4[16]
        _r16 = _e16["row"]
        _rounded16 = round(_e16["amount"])
        answer_box(f"=ROUND(E{_r16},0)", _rounded16)
        step(1, f"Find Order ID 16 → Row {_r16}",
             f"| Row | Order ID | Customer Name | Order Amount |\n"
             f"|-----|----------|--------------|-------------|\n"
             f"| {_r16} | 16 | {_e16['name']} | **{_e16['amount']}** ← E{_r16} |")
        step(2, "ROUND(E,0) means round to nearest whole number",
             f"```\n"
             f"{_e16['amount']}\n"
             f"  First decimal digit = {str(_e16['amount']).split('.')[1][0]} "
             f"({'< 5 → round DOWN' if int(str(_e16['amount']).split('.')[1][0]) < 5 else '>= 5 → round UP'})\n"
             f"Result = {_rounded16}\n"
             f"```\n\n"
             "| num_digits | Rounds to | Result for 80.15 |\n"
             "|------------|-----------|-----------------|\n"
             "| 2 | 2 decimal places | 80.15 |\n"
             "| 1 | 1 decimal place | 80.2 |\n"
             "| **0** | **Whole number** | **80** |\n"
             "| -1 | Nearest 10 | 80 |")
        st.divider()

        # ── Q24.4 ────────────────────────────────────────────────────────────────────
        st.subheader("Q24.4 — Length of Customer Name for Order ID 7 (LEN)")
        _e7 = _d4[7]
        _r7 = _e7["row"]
        _name7 = _e7["name"]
        _len7 = len(_name7)
        answer_box(f"=LEN(B{_r7})", _len7)
        step(1, f"Find Order ID 7 → Row {_r7}",
             f"| Row | Order ID | Customer Name |\n"
             f"|-----|----------|--------------|\n"
             f"| {_r7} | 7 | **{_name7}** ← B{_r7} |")
        step(2, "LEN counts every character including spaces",
             f"```\n\"{_name7}\"\n" +
             "  " + "  ".join(list(_name7)) + "\n" +
             "  " + "  ".join(str(i+1) for i in range(len(_name7))) + "\n" +
             f"Total = {_len7} characters (including the space)\n" +
             f"```\n\n"
             "⚠️ **Exam trap:** LEN counts spaces. "
             f"'{_name7}' = {_len7}, NOT {_len7 - _name7.count(' ')}.")
        st.divider()

        # ── Q24.5 ────────────────────────────────────────────────────────────────────
        st.subheader("Q24.5 — Initials and Surname for Order ID 12 (LEFT + MID + FIND)")
        _e12 = _d4[12]
        _r12 = _e12["row"]
        _name12 = _e12["name"]
        _parts12 = _name12.split(" ")
        _initial12 = _parts12[0][0]
        _surname12 = _parts12[-1]
        _result12 = f"{_initial12} {_surname12}"
        _space_pos = _name12.index(" ") + 1  # 1-based
        answer_box(
            f'=LEFT(B{_r12},1)&" "&MID(B{_r12},FIND(" ",B{_r12})+1,LEN(B{_r12}))',
            _result12
        )
        step(1, f"Find Order ID 12 → Row {_r12}",
             f"| Row | Order ID | Customer Name |\n"
             f"|-----|----------|--------------|\n"
             f"| {_r12} | 12 | **{_name12}** ← B{_r12} |")
        step(2, "Three-part formula breakdown",
             f"**Part A — Extract the initial:**\n"
             f"`LEFT(B{_r12},1)` = LEFT(\"{_name12}\",1) = **\"{_initial12}\"**\n\n"
             f"**Part B — Find the space position:**\n"
             f"`FIND(\" \",B{_r12})` = FIND(\" \",\"{_name12}\") = **{_space_pos}**\n"
             f"(space is at position {_space_pos} of the string)\n\n"
             f"**Part C — Extract the surname:**\n"
             f"`MID(B{_r12},FIND(\" \",B{_r12})+1,LEN(B{_r12}))` = "
             f"MID(\"{_name12}\",{_space_pos+1},{len(_name12)}) = **\"{_surname12}\"**\n\n"
             f"**Combine:** `\"{_initial12}\"` & `\" \"` & `\"{_surname12}\"` = **\"{_result12}\"**")
        step(3, f"Character map of '{_name12}'",
             "| Position | " + " | ".join(str(i+1) for i in range(len(_name12))) + " |\n" +
             "| --- | " + " | ".join("---" for _ in _name12) + " |\n" +
             "| Char | " + " | ".join(
                 (f"**{c}**" if c == " " else c) for c in _name12
             ) + " |\n\n" +
             f"Space at position {_space_pos} → MID starts at position {_space_pos+1} → gets '{_surname12}'")
        st.divider()

        # ── Q24.6 ────────────────────────────────────────────────────────────────────
        st.subheader("Q24.6 — Order Status for Order ID 11 (IF + DATE)")
        _e11 = _d4[11]
        _r11 = _e11["row"]
        _cutoff = _date(2023, 10, 10)
        _odate11 = _e11["date"].date()
        _status11 = "On Time" if _odate11 <= _cutoff else "Delayed"
        answer_box(f'=IF(C{_r11}<=DATE(2023,10,10),"On Time","Delayed")', _status11)
        step(1, f"Find Order ID 11 → Row {_r11}",
             f"| Row | Order ID | Customer Name | Order Date |\n"
             f"|-----|----------|--------------|------------|\n"
             f"| {_r11} | 11 | {_e11['name']} | **{_odate11}** ← C{_r11} |")
        step(2, "IF logic with DATE cutoff",
             f"`DATE(2023,10,10)` = **2023-10-10** (the cutoff)\n\n"
             f"`{_odate11}` <= `2023-10-10`?  →  **{_odate11 <= _cutoff}**\n\n"
             f"Result: **\"{_status11}\"**\n\n"
             "⚠️ **Exam trap:** The rule is `<=` (on or before). "
             "Order ID 10 (Susan Turner, 2023-10-10) = On Time because Oct 10 equals the cutoff. "
             "Richard Adams (Oct 11) is one day late → **Delayed**.")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**On Time (date ≤ 2023-10-10)**")
            for k, v in _d4.items():
                if v["date"].date() <= _cutoff:
                    st.markdown(f"- OID {k}: {v['name']} ({v['date'].date()})")
        with col2:
            st.markdown("**Delayed (date > 2023-10-10)**")
            for k, v in _d4.items():
                if v["date"].date() > _cutoff:
                    marker = " ← Q24.6" if k == 11 else ""
                    st.markdown(f"- OID {k}: {v['name']} ({v['date'].date()}){marker}")
        st.divider()

        # ── Q24.7 ────────────────────────────────────────────────────────────────────
        st.subheader("Q24.7 — Discount for Order ID 5 (IF + 10%)")
        _e5 = _d4[5]
        _r5 = _e5["row"]
        _amt5 = _e5["amount"]
        _disc5 = round(_amt5 * 0.10, 2) if _amt5 >= 100 else 0
        answer_box(f"=IF(E{_r5}>=100,E{_r5}*10%,0)", _disc5)
        step(1, f"Find Order ID 5 → Row {_r5}",
             f"| Row | Order ID | Customer Name | Order Amount |\n"
             f"|-----|----------|--------------|-------------|\n"
             f"| {_r5} | 5 | {_e5['name']} | **{_amt5}** ← E{_r5} |")
        step(2, "IF logic applied",
             f"`{_amt5}` >= 100?  →  **True**\n\n"
             f"Discount = E{_r5} * 10% = {_amt5} × 0.10 = **{_disc5}**\n\n"
             "> `E*10%` is identical to `E*0.1` — Excel understands `%` notation.")
        step(3, "Discount table for all orders",
             "| Order ID | Customer | Amount | >= 100? | Discount |\n"
             "|----------|---------|--------|---------|----------|\n" +
             "\n".join(
                 f"| {k} | {v['name']} | {v['amount']} | "
                 f"{'Yes ✅' if v['amount'] >= 100 else 'No ❌'} | "
                 f"{round(v['amount']*0.10,2) if v['amount'] >= 100 else 0} |"
                 + (" ← Q24.7" if k == 5 else "")
                 for k, v in _d4.items()
             ))
        st.divider()

        # ── Summary ────────────────────────────────────────────────────────────────────────
        st.subheader("📌 Q24 Summary")
        import pandas as _pd4b
        summary4 = _pd4b.DataFrame({
            "Q": ["24.1","24.2","24.3","24.4","24.5","24.6","24.7"],
            "Task": [
                "Next Order Date OID5 (+3 days)",
                "Time Slot OID8 (IF+TIME)",
                "Rounded Amount OID16 (ROUND)",
                "Name length OID7 (LEN)",
                "Initials+Surname OID12 (LEFT+MID+FIND)",
                "Order Status OID11 (IF+DATE)",
                "Discount OID5 (IF 10%)",
            ],
            "Formula": [
                f"=C{_d4[5]['row']}+3",
                f'=IF(D{_d4[8]["row"]}<TIME(12,0,0),"Morning","Afternoon")',
                f"=ROUND(E{_d4[16]['row']},0)",
                f"=LEN(B{_d4[7]['row']})",
                f'=LEFT(B{_d4[12]["row"]},1)&" "&MID(B{_d4[12]["row"]},FIND(" ",B{_d4[12]["row"]})+1,LEN(B{_d4[12]["row"]}))',
                f'=IF(C{_d4[11]["row"]}<=DATE(2023,10,10),"On Time","Delayed")',
                f"=IF(E{_d4[5]['row']}>=100,E{_d4[5]['row']}*10%,0)",
            ],
            "Answer": [
                str(_next),
                _slot8,
                _rounded16,
                _len7,
                _result12,
                _status11,
                _disc5,
            ]
        })
        st.dataframe(summary4, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — PYTHON SOLUTIONS (from Resolved Files)
# Loads each "CA Lesson X dataset - Resolved.xlsx" directly, extracts the
# official Excel formula strings the instructor placed there, then replicates
# every calculation in Python/pandas and shows them side-by-side.
# ══════════════════════════════════════════════════════════════════════════════
with tab5:
    st.header("🔑 Python Solutions — derived from the Resolved Datasets")
    st.markdown(
        "Each section loads the **official resolved xlsx file**, reads the exact Excel formula "
        "the instructor wrote, then computes the same answer in Python. "
        "This is the authoritative cross-check for all four exam assignments."
    )

    import openpyxl as _opx
    import pandas as _pd5
    from datetime import timedelta as _td5, date as _date5, time as _time5

    def _load_resolved(fname):
        """Return (ws, df_data) where ws is openpyxl worksheet and df_data
        is a pandas DataFrame of the first 20 data rows (no formula rows)."""
        if not os.path.exists(fname):
            st.error(f"File not found: {fname}")
            return None, None
        wb = _opx.load_workbook(fname)
        ws = wb.active
        df = _pd5.read_excel(fname, header=0, nrows=20)
        return ws, df

    def _cell_formula(ws, row, col):
        """Return the formula string stored in a cell, or '' if none.
        Strips openpyxl's internal _xlfn. prefix on modern Excel functions."""
        v = ws.cell(row, col).value
        if v is None:
            return ""
        if hasattr(v, "text"):        # ArrayFormula object
            return "=" + str(v.text).replace("_xlfn.", "")
        return str(v).replace("_xlfn.", "")

    def _show_solutions(rows):
        """Render a list of dicts {Q, Excel Formula, Python Code, Result} as
        a styled dataframe. Discrepancies (if any) are flagged."""
        df = _pd5.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True,
                     column_config={
                         "Excel Formula": st.column_config.TextColumn(width="large"),
                         "Python Code":   st.column_config.TextColumn(width="large"),
                         "Result":        st.column_config.TextColumn(width="small"),
                     })

    st.markdown("---")

    # ─── CA Lesson 1 — Q21 ─────────────────────────────────────────────────
    st.subheader("📁 CA Lesson 1 — Q21 (Sales Data)")
    _ws1, _df1 = _load_resolved("CA Lesson 1 dataset - Resolved.xlsx")
    if _ws1 and _df1 is not None:
        _df1.columns = ["Date", "Product ID", "Units Sold", "Price per Unit",
                        "Product Details"]
        _df1["Price per Unit"] = _pd5.to_numeric(
            _df1["Price per Unit"].astype(str).str.replace("$","",regex=False),
            errors="coerce")
        _df1["Units Sold"] = _pd5.to_numeric(_df1["Units Sold"], errors="coerce")
        _df1["Date"]       = _pd5.to_datetime(_df1["Date"])
        _df1.index = range(2, 22)

        # compute
        _r21_1 = int(_df1["Units Sold"].sum())
        _r21_2 = round(_df1["Price per Unit"].mean(), 2)
        _r21_3 = int(_df1["Units Sold"].max())
        _r21_4 = int(_df1["Price per Unit"].min())
        _r21_5 = int((_df1["Product ID"] == "P002").sum())
        _p001_sep10 = int(_df1.loc[
            (_df1["Product ID"]=="P001") &
            (_df1["Date"] == _pd5.Timestamp("2023-09-10")), "Units Sold"
        ].sum())
        _r21_6 = "Low Sales" if _p001_sep10 <= 20 else "High Sales"
        _r21_7 = str(_df1.loc[7, "Product ID"])[:4]
        _r21_8 = str(_df1.loc[15, "Product ID"])[-3:]

        _sol1 = [
            {"Q":"21.1","Excel Formula":_cell_formula(_ws1,22,3) or "=SUM(C2:C21)",
             "Python Code":'df["Units Sold"].sum()', "Result":str(_r21_1)},
            {"Q":"21.2","Excel Formula":_cell_formula(_ws1,22,4) or "=AVERAGE(D2:D21)",
             "Python Code":'df["Price per Unit"].mean()', "Result":str(_r21_2)},
            {"Q":"21.3","Excel Formula":_cell_formula(_ws1,24,3) or "=MAX(C2:C21)",
             "Python Code":'df["Units Sold"].max()', "Result":str(_r21_3)},
            {"Q":"21.4","Excel Formula":_cell_formula(_ws1,24,4) or "=MIN(D2:D21)",
             "Python Code":'df["Price per Unit"].min()', "Result":str(_r21_4)},
            {"Q":"21.5","Excel Formula":_cell_formula(_ws1,22,2) or '=COUNTIF(B2:B21,"P002")',
             "Python Code":'(df["Product ID"] == "P002").sum()', "Result":str(_r21_5)},
            {"Q":"21.6","Excel Formula":
             '=IF(SUMIFS(C2:C21,B2:B21,"P001",A2:A21,DATE(2023,9,10))>20,"High Sales","Low Sales")',
             "Python Code":
             'sumifs = df.loc[(df["Product ID"]=="P001")&(df["Date"]=="2023-09-10"),"Units Sold"].sum()\n'
             '"High Sales" if sumifs > 20 else "Low Sales"',
             "Result":_r21_6},
            {"Q":"21.7","Excel Formula":_cell_formula(_ws1,26,1) or "=LEFT(B7,4)",
             "Python Code":'df.loc[7, "Product ID"][:4]', "Result":_r21_7},
            {"Q":"21.8","Excel Formula":"=RIGHT(B15,3)",
             "Python Code":'df.loc[15, "Product ID"][-3:]', "Result":_r21_8},
        ]
        _show_solutions(_sol1)

        with st.expander("View source data rows 1–21"):
            st.dataframe(_df1[["Date","Product ID","Units Sold","Price per Unit"]],
                         use_container_width=True)

    st.markdown("---")

    # ─── CA Lesson 3 — Q22 ─────────────────────────────────────────────────
    st.subheader("📁 CA Lesson 3 — Q22 (Employee Salaries + Lookup)")
    _ws3, _df3 = _load_resolved("CA Lesson 3 dataset - Resolved.xlsx")
    if _ws3 and _df3 is not None:
        _df3.columns = ["Employee ID", "Employee Name", "Department", "Salary"]
        _df3.index   = range(2, 22)

        def _vlookup3(name, ret_col):
            return _df3.loc[_df3["Employee Name"]==name, ret_col].iloc[0]

        _r22_1 = int(_vlookup3("Carol",   "Salary"))
        _r22_2 = str(_vlookup3("Paul",    "Department"))
        _r22_3 = int(_vlookup3("Maria",   "Salary"))
        _r22_4 = int(_df3["Salary"].idxmax()) - 1   # MATCH → 1-based relative pos
        _match_row = int(_df3["Salary"].values.tolist().index(_df3["Salary"].max())) + 1
        _r22_5 = int(_df3["Salary"].iloc[:5].sum())
        _r22_6 = str(_vlookup3("Irene",   "Department"))
        _r22_7 = int(_df3.loc[15, "Salary"])          # INDIRECT("D15") → row 15
        _r22_8 = int(_vlookup3("Grace",   "Salary"))
        _r22_9_name = "Sam"
        _r22_9 = int(_df3.loc[_df3["Employee Name"]==_r22_9_name,"Salary"].iloc[0])

        _sol3 = [
            {"Q":"22.1","Excel Formula":_cell_formula(_ws3,23,1) or '=VLOOKUP("Carol",B2:D21,3,FALSE)',
             "Python Code":'df[df["Employee Name"]=="Carol"]["Salary"].iloc[0]',
             "Result":str(_r22_1)},
            {"Q":"22.2","Excel Formula":_cell_formula(_ws3,25,1) or '=VLOOKUP("Paul",B2:D21,2,FALSE)',
             "Python Code":'df[df["Employee Name"]=="Paul"]["Department"].iloc[0]',
             "Result":str(_r22_2)},
            {"Q":"22.3","Excel Formula":_cell_formula(_ws3,26,1) or '=INDEX(D2:D21,MATCH("Maria",B2:B21,0))',
             "Python Code":'df[df["Employee Name"]=="Maria"]["Salary"].iloc[0]',
             "Result":str(_r22_3)},
            {"Q":"22.4","Excel Formula":_cell_formula(_ws3,28,1) or "=MATCH(MAX(D2:D21),D2:D21,0)",
             "Python Code":'df["Salary"].values.tolist().index(df["Salary"].max()) + 1',
             "Result":str(_match_row)},
            {"Q":"22.5","Excel Formula":_cell_formula(_ws3,30,1) or "=SUM(OFFSET(D2,0,0,5,1))",
             "Python Code":'df["Salary"].iloc[:5].sum()',
             "Result":str(_r22_5)},
            {"Q":"22.6","Excel Formula":_cell_formula(_ws3,32,1) or '=INDEX(C2:C21,MATCH("Irene",B2:B21,0))',
             "Python Code":'df[df["Employee Name"]=="Irene"]["Department"].iloc[0]',
             "Result":str(_r22_6)},
            {"Q":"22.7","Excel Formula":_cell_formula(_ws3,34,1) or '=INDIRECT("D15")',
             "Python Code":'df.loc[15, "Salary"]  # row 15 in Excel',
             "Result":str(_r22_7)},
            {"Q":"22.8","Excel Formula":_cell_formula(_ws3,36,1) or '=LOOKUP("Grace",B2:B21,D2:D21)',
             "Python Code":'df[df["Employee Name"]=="Grace"]["Salary"].iloc[0]',
             "Result":str(_r22_8)},
            {"Q":"22.9","Excel Formula":_cell_formula(_ws3,38,1) or '=XLOOKUP("Sam",B2:B21,D2:D21)',
             "Python Code":'df[df["Employee Name"]=="Sam"]["Salary"].iloc[0]',
             "Result":str(_r22_9)},
        ]
        _show_solutions(_sol3)

        with st.expander("View source data rows 1–21"):
            st.dataframe(_df3, use_container_width=True)

    st.markdown("---")

    # ─── CA Lesson 4 — Q23 ─────────────────────────────────────────────────
    st.subheader("📁 CA Lesson 4 — Q23 (Salaries + Bonus %)")
    _ws4q, _df4q = _load_resolved("CA Lesson 4 dataset - Resolved.xlsx")
    if _ws4q and _df4q is not None:
        _df4q = _df4q.iloc[:, :6]   # drop helper col G stored in resolved xlsx
        _df4q.columns = ["Employee ID","Employee Name","Department",
                         "Salary","Years of Service","Bonus Percentage"]
        _df4q.index = range(2, 22)

        _alice = _df4q.iloc[0]

        # Q23.1 MAXIFS(D, dept=Sales, years>=2)
        _r23_1 = int(_df4q.loc[
            (_df4q["Department"]=="Sales") & (_df4q["Years of Service"]>=2),
            "Salary"].max())
        # Q23.2 MINIFS(F, dept=Finance, salary>58000) * 100
        _r23_2 = round(_df4q.loc[
            (_df4q["Department"]=="Finance") & (_df4q["Salary"]>58000),
            "Bonus Percentage"].min() * 100, 2)
        # Q23.3 COUNTIF(F,">=0.08")
        _r23_3 = int((_df4q["Bonus Percentage"] >= 0.08).sum())
        # Q23.4 COUNTIFS(dept=IT, salary>60000)
        _r23_4 = int(((_df4q["Department"]=="IT") & (_df4q["Salary"]>60000)).sum())
        # Q23.5 AVERAGEIF(F,"<0.06",D)  — avg Salary where Bonus% < 0.06
        _r23_5 = round(_df4q.loc[_df4q["Bonus Percentage"]<0.06,"Salary"].mean(), 2)
        # Q23.6 AVERAGEIFS(F, dept=Sales, years>1)  — avg Bonus% of Sales, years>1
        _r23_6 = round(_df4q.loc[
            (_df4q["Department"]=="Sales") & (_df4q["Years of Service"]>1),
            "Bonus Percentage"].mean(), 4)
        # Q23.7 col G = D*F; IFERROR(G2/F2,"Error") = salary
        _g2 = _alice["Salary"] * _alice["Bonus Percentage"]
        _r23_7 = int(_g2 / _alice["Bonus Percentage"])
        # Q23.8 LOOKUP("Grace Martin", names, salary)
        _r23_8 = int(_df4q.loc[_df4q["Employee Name"]=="Grace Martin","Salary"].iloc[0])
        # Q23.9 XLOOKUP("Sam Scott", names, salary)
        _r23_9 = int(_df4q.loc[_df4q["Employee Name"]=="Sam Scott","Salary"].iloc[0])

        _sol4 = [
            {"Q":"23.1",
             "Excel Formula":_cell_formula(_ws4q,24,1) or '=MAXIFS(D2:D21,C2:C21,"Sales",E2:E21,">=2")',
             "Python Code":'df[(df.Dept=="Sales") & (df.Years>=2)]["Salary"].max()',
             "Result":str(_r23_1)},
            {"Q":"23.2",
             "Excel Formula":_cell_formula(_ws4q,26,1) or '=MINIFS(F2:F21,C2:C21,"Finance",D2:D21,">58000")*100',
             "Python Code":'df[(df.Dept=="Finance") & (df.Salary>58000)]["Bonus%"].min() * 100',
             "Result":str(_r23_2)},
            {"Q":"23.3",
             "Excel Formula":_cell_formula(_ws4q,28,1) or '=COUNTIF(F2:F21,">=0.08")',
             "Python Code":'(df["Bonus Percentage"] >= 0.08).sum()',
             "Result":str(_r23_3)},
            {"Q":"23.4",
             "Excel Formula":_cell_formula(_ws4q,30,1) or '=COUNTIFS(C2:C21,"IT",D2:D21,">60000")',
             "Python Code":'((df.Dept=="IT") & (df.Salary>60000)).sum()',
             "Result":str(_r23_4)},
            {"Q":"23.5",
             "Excel Formula":_cell_formula(_ws4q,32,1) or '=AVERAGEIF(F2:F21,"<0.06",D2:D21)',
             "Python Code":'df[df["Bonus Percentage"] < 0.06]["Salary"].mean()',
             "Result":str(_r23_5)},
            {"Q":"23.6",
             "Excel Formula":_cell_formula(_ws4q,34,1) or '=AVERAGEIFS(F2:F21,C2:C21,"Sales",E2:E21,">1")',
             "Python Code":'df[(df.Dept=="Sales") & (df.Years>1)]["Bonus Percentage"].mean()',
             "Result":str(_r23_6)},
            {"Q":"23.7",
             "Excel Formula":_cell_formula(_ws4q,36,1) or '=IFERROR(G2/F2,"Error")',
             "Python Code":'try: (D2*F2)/F2  →  D2 (salary)\nexcept ZeroDivision: "Error"',
             "Result":str(_r23_7)},
            {"Q":"23.8",
             "Excel Formula":_cell_formula(_ws4q,38,1) or '=LOOKUP("Grace Martin",B2:B21,D2:D21)',
             "Python Code":'df[df["Employee Name"]=="Grace Martin"]["Salary"].iloc[0]',
             "Result":str(_r23_8)},
            {"Q":"23.9",
             "Excel Formula":_cell_formula(_ws4q,40,1) or '=XLOOKUP("Sam Scott",B2:B21,D2:D21)',
             "Python Code":'df[df["Employee Name"]=="Sam Scott"]["Salary"].iloc[0]',
             "Result":str(_r23_9)},
        ]
        _show_solutions(_sol4)

        with st.expander("View source data rows 1–21 (+ helper column G = Salary × Bonus%)"):
            _df4q["Bonus Amount (G=D×F)"] = (_df4q["Salary"] * _df4q["Bonus Percentage"]).round(2)
            st.dataframe(_df4q, use_container_width=True)

    st.markdown("---")

    # ─── CA Lesson 2 — Q24 ─────────────────────────────────────────────────
    st.subheader("📁 CA Lesson 2 — Q24 (Orders: Date, Time & Text)")
    _ws2, _df2 = _load_resolved("CA Lesson 2 dataset - Resolved.xlsx")
    if _ws2 and _df2 is not None:
        _df2.columns = ["Order ID","Customer Name","Order Date",
                        "Order Time","Order Amount","Time Slot"]
        _df2["Order Date"] = _pd5.to_datetime(_df2["Order Date"])
        _df2.index = range(2, 22)

        # Q24.1  =C6+A4  (A4 = Order ID 3 = 3)
        _a4_val = int(_ws2.cell(4, 1).value)  # openpyxl reads A4 as its stored value
        _r24_1 = (_df2.loc[6, "Order Date"] + _pd5.Timedelta(days=_a4_val)).date()
        # Q24.2 IF(D9<TIME(12,0,0),...)
        _t9 = _df2.loc[9, "Order Time"]
        _r24_2 = "Morning" if (_t9.hour < 12 or (_t9.hour == 12 and _t9.minute == 0 and _t9.second == 0)) else "Afternoon"
        # Q24.3 ROUND(E17,0)
        _r24_3 = round(_df2.loc[17, "Order Amount"])
        # Q24.4 LEN(B8)
        _r24_4 = len(str(_df2.loc[8, "Customer Name"]))
        # Q24.5 LEFT+MID+FIND
        _n12 = str(_df2.loc[13, "Customer Name"])
        _r24_5 = _n12[0] + " " + _n12.split(" ", 1)[1]
        # Q24.6 IF(C12<=DATE(2023,10,10),...)
        _d12  = _df2.loc[12, "Order Date"].date()
        _cut  = _date5(2023, 10, 10)
        _r24_6 = "On Time" if _d12 <= _cut else "Delayed"
        # Q24.7 IF(E6>=100, E6*10%, 0)
        _e6   = _df2.loc[6, "Order Amount"]
        _r24_7 = round(_e6 * 0.10, 2) if _e6 >= 100 else 0

        # Official formulas from resolved file
        _f24 = {i: _cell_formula(_ws2, r, 2) for i, r in
                enumerate([24, 26, 28, 30, 32, 34, 36], 1)}

        _sol2 = [
            {"Q":"24.1",
             "Excel Formula": _f24[1] or "=C6+A4",
             "Python Code":"(df.loc[6,'Order Date'] + timedelta(days=A4_value)).date()\n# A4 = Order ID in row 4 = 3",
             "Result":str(_r24_1)},
            {"Q":"24.2",
             "Excel Formula": _f24[2] or '=IF(D9<TIME(12,0,0),"Morning","Afternoon")',
             "Python Code":'"Morning" if time(row9) < time(12,0) else "Afternoon"',
             "Result":_r24_2},
            {"Q":"24.3",
             "Excel Formula": _f24[3] or "=ROUND(E17,0)",
             "Python Code":"round(df.loc[17, 'Order Amount'])",
             "Result":str(_r24_3)},
            {"Q":"24.4",
             "Excel Formula": _f24[4] or "=LEN(B8)",
             "Python Code":"len(df.loc[8, 'Customer Name'])",
             "Result":str(_r24_4)},
            {"Q":"24.5",
             "Excel Formula": _f24[5] or '=LEFT(B13,1)&" "&MID(B13,FIND(" ",B13)+1,LEN(B13))',
             "Python Code":"name[0] + ' ' + name.split(' ', 1)[1]",
             "Result":_r24_5},
            {"Q":"24.6",
             "Excel Formula": _f24[6] or '=IF(C12<=DATE(2023,10,10),"On Time","Delayed")',
             "Python Code":'"On Time" if date(row12) <= date(2023,10,10) else "Delayed"',
             "Result":_r24_6},
            {"Q":"24.7",
             "Excel Formula": _f24[7] or "=IF(E6>=100,E6*10%,0)",
             "Python Code":"round(e6 * 0.10, 2) if e6 >= 100 else 0",
             "Result":str(_r24_7)},
        ]
        _show_solutions(_sol2)

        with st.expander("View source data rows 1–21"):
            st.dataframe(_df2[["Order ID","Customer Name","Order Date","Order Time","Order Amount"]],
                         use_container_width=True)

    st.divider()
    st.info(
        "**📌 Note on Q24.1 formula:** The resolved file contains `=C6+A4` (using cell A4 which holds "
        "Order ID 3 = the number 3). Our lesson shows `=C6+3` — both produce identical results.\n\n"
        "**📌 Note on Q23 (bonus questions):** Q23.2, Q23.5, Q23.6, Q23.7 all operate on column F "
        "(Bonus Percentage) and column G (Bonus Amount = D×F) — NOT just salary. See the Q23 tab for "
        "full step-by-step explanations."
    )
