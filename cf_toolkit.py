"""
cf_toolkit.py  –  Reusable Conditional Formatting Toolkit
==========================================================
Course: FI1BBSF05 Spreadsheet Fundamentals – Lesson 1.4

HOW TO REUSE THIS SCRIPT
─────────────────────────
1. Change the CONFIG block at the top to point to YOUR file.
2. Set HEADER_ROW to the row number that contains your column headers.
3. Change COLUMNS to map friendly names to the actual Excel column letters.
4. Run:  python cf_toolkit.py
5. Open the OUTPUT file in Excel or LibreOffice Calc.

WHAT THIS SCRIPT CONTAINS
──────────────────────────
Five ready-to-call functions you can mix and match:

  apply_highlight(ws, cell_range, operator, values, bg_hex, fg_hex)
      → Highlight cells based on a condition  (Tasks 1, 2, 3)

  apply_color_scale(ws, cell_range, colours)
      → Gradient colour fill across the range  (Tasks 4, 5)

  apply_data_bars(ws, cell_range, colour)
      → Mini bar-chart inside the cells  (Task 6)

  apply_icon_set(ws, cell_range, style, thresholds)
      → Icons (arrows, traffic lights, etc.)  (Task 7)

  apply_formula_rule(ws, cell_range, formula, bg_hex, fg_hex)
      → Any custom COUNTIF / LARGE / SMALL formula  (Tasks 8, 9, 10)
"""

# ══════════════════════════════════════════════════════════════════
# ①  CONFIG  –  Edit this section to adapt the script to any file
# ══════════════════════════════════════════════════════════════════

SOURCE_FILE = "SPF+-+Sample+dataset+Activity+2.xlsx"   # ← your input file
OUTPUT_FILE = "CF_Output.xlsx"                         # ← output file name

HEADER_ROW  = 1   # Row number that contains column headers (usually 1)

# Map a friendly name  →  Excel column letter
# Change these to match your own spreadsheet columns.
COLUMNS = {
    "name":  "A",
    "math":  "B",
    "sci":   "C",
    "eng":   "D",
}

# Rules toggles – set False to skip a rule
RULES = {
    "task1_math_below_60":       True,
    "task2_sci_above_90":        True,
    "task3_eng_between_70_80":   True,
    "task4_math_color_scale_3":  True,
    "task5_sci_color_scale_2":   True,
    "task6_eng_data_bars":       True,
    "task7_math_icons":          True,
    "task8_duplicate_names":     True,
    "task9_top3_sci":            True,
    "task10_bottom3_math":       True,
}

# ══════════════════════════════════════════════════════════════════
# ②  IMPORTS
# ══════════════════════════════════════════════════════════════════
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    IconSetRule,
    FormulaRule,
)

# ══════════════════════════════════════════════════════════════════
# ③  HELPER UTILITIES
# ══════════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> PatternFill:
    """Return a solid PatternFill from a hex colour string (no #)."""
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _font(hex_color: str, bold: bool = True) -> Font:
    """Return a Font with the given hex colour."""
    return Font(color=hex_color, bold=bold)

def col_range(ws, col_letter: str, header_row: int) -> str:
    """
    Build an Excel range string like 'B2:B31' for the data in a column,
    automatically skipping the header row.

    Example:
        col_range(ws, "B", 1)  →  "B2:B31"
    """
    first = header_row + 1
    last  = ws.max_row
    return f"{col_letter}{first}:{col_letter}{last}"

# ══════════════════════════════════════════════════════════════════
# ④  REUSABLE FORMATTING FUNCTIONS
# ══════════════════════════════════════════════════════════════════

def apply_highlight(ws, cell_range: str, operator: str,
                    values: list, bg_hex: str, fg_hex: str = None):
    """
    Highlight cells that meet a simple condition.

    Parameters
    ----------
    ws          : the worksheet object
    cell_range  : e.g. "B2:B31"
    operator    : one of  "lessThan", "greaterThan", "between",
                          "equal", "notEqual", "greaterThanOrEqual",
                          "lessThanOrEqual"
    values      : list with one value ["60"]  or two for "between" ["70","80"]
    bg_hex      : background colour hex, e.g. "FF0000" for red
    fg_hex      : (optional) text colour hex

    Examples
    --------
    # Maths < 60 → red background
    apply_highlight(ws, "B2:B31", "lessThan", ["60"], "FF0000")

    # English between 70 and 80 → yellow
    apply_highlight(ws, "D2:D31", "between", ["70","80"], "FFFF00")
    """
    fmt_font = _font(fg_hex) if fg_hex else None
    rule = CellIsRule(operator=operator, formula=values,
                      fill=_fill(bg_hex), font=fmt_font)
    ws.conditional_formatting.add(cell_range, rule)
    print(f"  ✓  Highlight applied  [{cell_range}]  {operator}  {values}  → #{bg_hex}")


def apply_color_scale(ws, cell_range: str, colours: list):
    """
    Apply a 2-colour or 3-colour gradient scale.

    colours must be a list of 2 or 3 hex strings:
        2 colours  →  ["start_hex", "end_hex"]
        3 colours  →  ["start_hex", "mid_hex", "end_hex"]

    Examples
    --------
    # 3-colour: red → yellow → green
    apply_color_scale(ws, "B2:B31", ["FF0000", "FFFF00", "00FF00"])

    # 2-colour: pink → blue
    apply_color_scale(ws, "C2:C31", ["FFAAAA", "0000FF"])
    """
    if len(colours) == 3:
        rule = ColorScaleRule(
            start_type="min",        start_color=colours[0],
            mid_type="percentile",   mid_value=50, mid_color=colours[1],
            end_type="max",          end_color=colours[2],
        )
    elif len(colours) == 2:
        rule = ColorScaleRule(
            start_type="min", start_color=colours[0],
            end_type="max",   end_color=colours[1],
        )
    else:
        raise ValueError("colours must have 2 or 3 items")

    ws.conditional_formatting.add(cell_range, rule)
    print(f"  ✓  Colour scale applied  [{cell_range}]  {len(colours)}-colour")


def apply_data_bars(ws, cell_range: str, colour: str = "4472C4"):
    """
    Add data bars to a range.

    colour : hex for the bar colour (default: Excel blue "4472C4")

    Example
    -------
    apply_data_bars(ws, "D2:D31")           # blue bars
    apply_data_bars(ws, "D2:D31", "FF0000") # red bars
    """
    rule = DataBarRule(
        start_type="min", start_value=0,
        end_type="max",   end_value=100,
        color=colour,
    )
    ws.conditional_formatting.add(cell_range, rule)
    print(f"  ✓  Data bars applied  [{cell_range}]  colour=#{colour}")


def apply_icon_set(ws, cell_range: str,
                   style: str = "3Arrows",
                   thresholds: list = None):
    """
    Apply an icon set to a range.

    style      : "3Arrows", "3TrafficLights1", "3Stars", "5Rating", etc.
    thresholds : list of percent breakpoints.
                 Default for 3-icon sets → [0, 33, 67]
                 For 5-icon sets → [0, 20, 40, 60, 80]

    Example
    -------
    apply_icon_set(ws, "B2:B31", "3Arrows")
    apply_icon_set(ws, "B2:B31", "3TrafficLights1", [0, 40, 70])
    """
    if thresholds is None:
        n = int(style[0])   # first char is usually the icon count
        thresholds = [int(100 / n * i) for i in range(n)]

    rule = IconSetRule(
        icon_style=style,
        type="percent",
        values=thresholds,
        showValue=True,
    )
    ws.conditional_formatting.add(cell_range, rule)
    print(f"  ✓  Icon set applied  [{cell_range}]  style={style}  thresholds={thresholds}")


def apply_formula_rule(ws, cell_range: str, formula: str,
                       bg_hex: str, fg_hex: str = None):
    """
    Apply any custom formula-based conditional formatting rule.
    This is the most powerful and flexible function.

    formula : Excel formula as a string.
              • Always write it as if you're in the FIRST cell of the range.
              • Use $COL$ROW for absolute references (stay fixed as rule copies down).
              • Use COL+ROW  for relative references (adjust per row).

    Examples
    --------
    # Highlight duplicates in A2:A31
    apply_formula_rule(ws, "A2:A31",
                       "COUNTIF($A$2:$A$31,A2)>1",
                       "FFC7CE", "9C0006")

    # Top 3 values in C2:C31
    apply_formula_rule(ws, "C2:C31",
                       "C2>=LARGE($C$2:$C$31,3)",
                       "C6EFCE", "276221")

    # Bottom 3 values in B2:B31
    apply_formula_rule(ws, "B2:B31",
                       "B2<=SMALL($B$2:$B$31,3)",
                       "FFD7CE", "9C0006")
    """
    fmt_font = _font(fg_hex) if fg_hex else None
    rule = FormulaRule(formula=[formula], fill=_fill(bg_hex), font=fmt_font)
    ws.conditional_formatting.add(cell_range, rule)
    print(f"  ✓  Formula rule applied  [{cell_range}]  formula={formula}")


# ══════════════════════════════════════════════════════════════════
# ⑤  MAIN  –  Apply all 10 tasks using the functions above
# ══════════════════════════════════════════════════════════════════

def main():
    # ── Load ──────────────────────────────────────────────────────
    wb = load_workbook(SOURCE_FILE)
    ws = wb.active
    last_row = ws.max_row
    print(f"Loaded: '{ws.title}'   rows={last_row}  (header on row {HEADER_ROW})\n")

    # ── Build column ranges from CONFIG ───────────────────────────
    r = {name: col_range(ws, letter, HEADER_ROW)
         for name, letter in COLUMNS.items()}
    # e.g.  r["math"] = "B2:B31"

    # ── Demo: inject duplicate names so Task 8 is visible ─────────
    ws["A3"]  = "Student 1"
    ws["A10"] = "Student 5"
    print("Injected two duplicate names for Task 8 demo.\n")
    print("Applying conditional formatting rules:")

    # ── Task 1 ────────────────────────────────────────────────────
    if RULES["task1_math_below_60"]:
        apply_highlight(ws, r["math"], "lessThan", ["60"], "FF0000")

    # ── Task 2 ────────────────────────────────────────────────────
    if RULES["task2_sci_above_90"]:
        apply_highlight(ws, r["sci"], "greaterThan", ["90"], "00B050")

    # ── Task 3 ────────────────────────────────────────────────────
    if RULES["task3_eng_between_70_80"]:
        apply_highlight(ws, r["eng"], "between", ["70", "80"], "FFFF00")

    # ── Task 4 ────────────────────────────────────────────────────
    if RULES["task4_math_color_scale_3"]:
        apply_color_scale(ws, r["math"], ["FF0000", "FFFF00", "00FF00"])

    # ── Task 5 ────────────────────────────────────────────────────
    if RULES["task5_sci_color_scale_2"]:
        apply_color_scale(ws, r["sci"], ["FFAAAA", "0000FF"])

    # ── Task 6 ────────────────────────────────────────────────────
    if RULES["task6_eng_data_bars"]:
        apply_data_bars(ws, r["eng"])

    # ── Task 7 ────────────────────────────────────────────────────
    if RULES["task7_math_icons"]:
        apply_icon_set(ws, r["math"], "3Arrows")

    # ── Task 8 (uses absolute + relative refs in formula) ─────────
    name_col = COLUMNS["name"]
    name_rng = r["name"]
    first_data = HEADER_ROW + 1
    if RULES["task8_duplicate_names"]:
        apply_formula_rule(
            ws, name_rng,
            formula=f"COUNTIF(${name_col}${first_data}:${name_col}${last_row},{name_col}{first_data})>1",
            bg_hex="FFC7CE", fg_hex="9C0006",
        )

    # ── Task 9 ────────────────────────────────────────────────────
    sci_col = COLUMNS["sci"]
    if RULES["task9_top3_sci"]:
        apply_formula_rule(
            ws, r["sci"],
            formula=f"{sci_col}{first_data}>=LARGE(${sci_col}${first_data}:${sci_col}${last_row},3)",
            bg_hex="C6EFCE", fg_hex="276221",
        )

    # ── Task 10 ───────────────────────────────────────────────────
    math_col = COLUMNS["math"]
    if RULES["task10_bottom3_math"]:
        apply_formula_rule(
            ws, r["math"],
            formula=f"{math_col}{first_data}<=SMALL(${math_col}${first_data}:${math_col}${last_row},3)",
            bg_hex="FFD7CE", fg_hex="9C0006",
        )

    # ── Header styling ────────────────────────────────────────────
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)
    for cell in ws[HEADER_ROW]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    # ── Save ──────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"\n✓  Saved → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
