"""
SPF 1.4 Lesson Task — Data Validation & Conditional Formatting
Dataset: SPF-0104+Task+DATASET+-+Staff+Salaries.CSV

Question 1 — Data Validation on Age, Department, Salary
Question 2 — Conditional Formatting: Colour Scale, Data Bars, Icon Sets

Run:  python spf_0104_task.py
Output: SPF_0104_Task_Solution.xlsx
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
    CellIsRule, FormulaRule
)
from openpyxl.styles import Color

# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────
CSV_FILE   = "SPF-0104+Task+DATASET+-+Staff+Salaries.CSV"
OUTPUT     = "SPF_0104_Task_Solution.xlsx"

# Departments accepted by validation (covers all departments in the dataset)
VALID_DEPARTMENTS = ["Sales", "Marketing", "Finance", "Engineering", "HR", "IT"]

# Age validation: whole numbers 18–65
AGE_MIN, AGE_MAX = 18, 65

# Salary validation: decimal 30,000–100,000
SAL_MIN, SAL_MAX = 30000, 100000

# Icon-set threshold (equal to / above / below)
ICONSET_SALARY_THRESHOLD = 52000

# 3 new employee rows added to demonstrate dynamic CF (Question 2 modification requirement)
NEW_EMPLOYEES = [
    ("Alice Murphy",    23, "IT",          37500.00),
    ("Cameron Reid",    48, "Marketing",   98000.00),
    ("Jordan Blake",    35, "Finance",     52000.00),   # exactly at threshold
]

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
INVALID_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow for flagged cells
INVALID_FONT = Font(color="9C5700", bold=True)
GOOD_FILL    = PatternFill("solid", fgColor="E2EFDA")
NEW_FILL     = PatternFill("solid", fgColor="D9E1F2")   # light blue — new rows

THIN  = Side(style="thin",   color="BFBFBF")
THICK = Side(style="medium", color="1F4E79")
CELL_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

def header_style(cell):
    cell.fill   = HEADER_FILL
    cell.font   = HEADER_FONT
    cell.border = HEADER_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

def data_cell_style(cell):
    cell.border    = CELL_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ─────────────────────────────────────────────────────────────
# LOAD CSV
# ─────────────────────────────────────────────────────────────
def load_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append({
                "Name":       r["Name"],
                "Age":        int(r["Age"]),
                "Department": r["Department"],
                "Salary":     float(r["Salary"]),
            })
    return rows

# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def build_workbook(rows):
    wb = Workbook()

    # ── 1. CREATE VALIDATION HELPER SHEET (hidden dropdown source) ────────────
    ws_lists = wb.active
    ws_lists.title = "Lists"
    for i, dept in enumerate(VALID_DEPARTMENTS, start=1):
        ws_lists.cell(row=i, column=1, value=dept)
    ws_lists.sheet_state = "hidden"          # hide so it doesn't clutter the UI

    # ── 2. MAIN DATA SHEET ────────────────────────────────────────────────────
    ws = wb.create_sheet("Staff Salaries")
    wb.active = ws

    headers = ["Name", "Age", "Department", "Salary"]
    col_widths = [22, 8, 16, 14]

    # Write headers
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    # Combine original rows + new employee rows
    all_rows = [(r["Name"], r["Age"], r["Department"], r["Salary"]) for r in rows]
    new_start = len(all_rows) + 2          # first Excel row of new employees
    for emp in NEW_EMPLOYEES:
        all_rows.append(emp)

    # Write data
    for r_idx, (name, age, dept, salary) in enumerate(all_rows, start=2):
        is_new   = r_idx >= new_start
        age_bad  = not (AGE_MIN <= age <= AGE_MAX)
        sal_bad  = not (SAL_MIN <= salary <= SAL_MAX)
        dept_bad = dept not in VALID_DEPARTMENTS

        row_data = [name, age, dept, salary]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            data_cell_style(cell)

            # Highlight cells that violate validation rules (audit colour)
            if c_idx == 2 and age_bad:
                cell.fill = INVALID_FILL;  cell.font = INVALID_FONT
            elif c_idx == 3 and dept_bad:
                cell.fill = INVALID_FILL;  cell.font = INVALID_FONT
            elif c_idx == 4 and sal_bad:
                cell.fill = INVALID_FILL;  cell.font = INVALID_FONT
            elif is_new:
                cell.fill = NEW_FILL       # blue tint for new employee rows

            # Number format for salary column
            if c_idx == 4:
                cell.number_format = "#,##0.00"

    total_rows = len(all_rows) + 1   # +1 for header (rows 1..n+1)

    # ── 3. FREEZE PANE & AUTO-FILTER ─────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{total_rows}"

    # ── 4. DATA VALIDATION ────────────────────────────────────────────────────

    # --- Q1a: Age — whole numbers between 18 and 65 ---
    dv_age = DataValidation(
        type="whole",
        operator="between",
        formula1=str(AGE_MIN),
        formula2=str(AGE_MAX),
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Invalid Age",
        error=(
            f"Age must be a whole number between {AGE_MIN} and {AGE_MAX}.\n"
            "Please enter a valid age."
        ),
        showInputMessage=True,
        promptTitle="Age Required",
        prompt=f"Enter a whole number between {AGE_MIN} and {AGE_MAX}.",
    )
    dv_age.sqref = f"B2:B{total_rows}"
    ws.add_data_validation(dv_age)

    # --- Q1b: Department — dropdown list ---
    dept_source = '"' + ",".join(VALID_DEPARTMENTS) + '"'
    dv_dept = DataValidation(
        type="list",
        formula1=dept_source,
        allow_blank=False,
        showDropDown=False,       # False = SHOW the dropdown arrow
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Invalid Department",
        error=(
            "Department must be one of:\n"
            + ", ".join(VALID_DEPARTMENTS)
        ),
        showInputMessage=True,
        promptTitle="Select Department",
        prompt="Choose from: " + ", ".join(VALID_DEPARTMENTS),
    )
    dv_dept.sqref = f"C2:C{total_rows}"
    ws.add_data_validation(dv_dept)

    # --- Q1c: Salary — decimal 30,000 to 100,000 ---
    dv_salary = DataValidation(
        type="decimal",
        operator="between",
        formula1=str(float(SAL_MIN)),
        formula2=str(float(SAL_MAX)),
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Invalid Salary",
        error=(
            f"Salary must be a decimal number between "
            f"{SAL_MIN:,} and {SAL_MAX:,}.\n"
            "Maximum two decimal places allowed."
        ),
        showInputMessage=True,
        promptTitle="Salary",
        prompt=f"Enter salary between {SAL_MIN:,} and {SAL_MAX:,} (max 2 decimal places).",
    )
    dv_salary.sqref = f"D2:D{total_rows}"
    ws.add_data_validation(dv_salary)

    # --- Q1c (extra): Custom formula for max 2 decimal places ---
    # =INT(D2*100)=D2*100  → TRUE only when the value has ≤ 2 decimal places
    dv_salary_dp = DataValidation(
        type="custom",
        formula1="=INT(D2*100)=D2*100",
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorStyle="warning",    # Warning (not Stop) so it doesn't block correct entries
        errorTitle="Too Many Decimal Places",
        error="Salary should have at most 2 decimal places (e.g. 55000.50).",
        showInputMessage=False,
    )
    dv_salary_dp.sqref = f"D2:D{total_rows}"
    ws.add_data_validation(dv_salary_dp)

    # ── 5. CONDITIONAL FORMATTING ─────────────────────────────────────────────
    data_range_salary = f"D2:D{total_rows}"
    data_range_age    = f"B2:B{total_rows}"

    # --- Q2a: COLOUR SCALE on Salary (white → dark blue for low → high) ---
    ws.conditional_formatting.add(
        data_range_salary,
        ColorScaleRule(
            start_type="min",  start_color="FFFFFF",   # low  → white
            mid_type="percentile", mid_value=50, mid_color="4472C4",   # mid   → medium blue
            end_type="max",    end_color="1F3864",      # high → dark navy
        ),
    )

    # --- Q2b: DATA BARS on Age ---
    ws.conditional_formatting.add(
        data_range_age,
        DataBarRule(
            start_type="min", start_value=None,
            end_type="max",   end_value=None,
            color="638EC6",   # blue data bars — matches Excel default gradient blue
            showValue=True,
        ),
    )

    # --- Q2c: ICON SET on Salary (vs threshold 52000) ---
    #   Red down arrow   : salary < 52000
    #   Yellow side arrow: salary = 52000   (between 52000 and 52001)
    #   Green up arrow   : salary > 52000   (>= 52001)
    ws.conditional_formatting.add(
        data_range_salary,
        IconSetRule(
            icon_style="3Arrows",
            type="num",
            values=[0, ICONSET_SALARY_THRESHOLD, ICONSET_SALARY_THRESHOLD + 1],
            showValue=True,
            reverse=False,
        ),
    )

    # ── 6. VALIDATION LEGEND SHEET ────────────────────────────────────────────
    ws_legend = wb.create_sheet("Legend")
    ws_legend.column_dimensions["A"].width = 28
    ws_legend.column_dimensions["B"].width = 50

    legend_header_fill = PatternFill("solid", fgColor="1F4E79")
    legend_header_font = Font(bold=True, color="FFFFFF")

    legend_entries = [
        ("SECTION", "Description"),
        ("── DATA VALIDATION ──", ""),
        ("Age (Column B)",
         f"Whole numbers only, between {AGE_MIN} and {AGE_MAX}. Error Style: STOP."),
        ("Department (Column C)",
         "List: " + ", ".join(VALID_DEPARTMENTS) + ". Error Style: STOP."),
        ("Salary (Column D) — Range",
         f"Decimal between {SAL_MIN:,} and {SAL_MAX:,}. Error Style: STOP."),
        ("Salary (Column D) — Decimals",
         "Custom formula =INT(D*100)=D*100 limits to max 2 decimal places. Error Style: WARNING."),
        ("── CELL AUDIT COLOURS ──", ""),
        ("Yellow cell highlight",
         "Existing row violates a validation rule (Age out of range, bad dept, bad salary)."),
        ("Blue cell highlight",
         "Newly added employee rows (rows added to show dynamic CF)."),
        ("── CONDITIONAL FORMATTING ──", ""),
        ("Colour Scale — Salary",
         "White (lowest) → Medium Blue (median) → Dark Navy (highest). Shows salary distribution."),
        ("Data Bars — Age (blue)",
         "Bar length proportional to age value. Blue gradient (Excel default colour). Easy visual ranking."),
        (f"Icon Set — Salary vs {ICONSET_SALARY_THRESHOLD:,}",
         f"🔼 Green up arrow = above {ICONSET_SALARY_THRESHOLD:,}  |  "
         f"➡ Yellow side arrow = exactly {ICONSET_SALARY_THRESHOLD:,}  |  "
         f"🔽 Red down arrow = below {ICONSET_SALARY_THRESHOLD:,}."),
    ]

    for row_i, (label, desc) in enumerate(legend_entries, start=1):
        ca = ws_legend.cell(row=row_i, column=1, value=label)
        cb = ws_legend.cell(row=row_i, column=2, value=desc)
        if row_i == 1:
            ca.fill = legend_header_fill;  ca.font = legend_header_font
            cb.fill = legend_header_fill;  cb.font = legend_header_font
        elif label.startswith("──"):
            ca.font = Font(bold=True, color="1F4E79")
            cb.font = Font(italic=True, color="595959")
        ws_legend.row_dimensions[row_i].height = 18
        ca.alignment = Alignment(vertical="center")
        cb.alignment = Alignment(wrap_text=True, vertical="center")

    # ── 7. VALIDATION REPORT SHEET ────────────────────────────────────────────
    ws_report = wb.create_sheet("Validation Report")
    ws_report.column_dimensions["A"].width = 5
    ws_report.column_dimensions["B"].width = 22
    ws_report.column_dimensions["C"].width = 8
    ws_report.column_dimensions["D"].width = 16
    ws_report.column_dimensions["E"].width = 14
    ws_report.column_dimensions["F"].width = 40

    rpt_headers = ["Row", "Name", "Age", "Department", "Salary", "Validation Issues"]
    for col, h in enumerate(rpt_headers, start=1):
        cell = ws_report.cell(row=1, column=col, value=h)
        header_style(cell)
        ws_report.column_dimensions[get_column_letter(col)].width = [5, 22, 8, 16, 14, 40][col-1]

    report_row = 2
    violation_count = 0

    for r_idx, (name, age, dept, salary) in enumerate(all_rows, start=2):
        issues = []
        if not (AGE_MIN <= age <= AGE_MAX):
            issues.append(f"Age {age} outside {AGE_MIN}–{AGE_MAX}")
        if dept not in VALID_DEPARTMENTS:
            issues.append(f"Department '{dept}' not in approved list")
        if not (SAL_MIN <= salary <= SAL_MAX):
            issues.append(f"Salary {salary:,.2f} outside {SAL_MIN:,}–{SAL_MAX:,}")

        if issues:
            violation_count += 1
            row_cells = [r_idx, name, age, dept, salary, " | ".join(issues)]
            for col, val in enumerate(row_cells, start=1):
                cell = ws_report.cell(row=report_row, column=col, value=val)
                data_cell_style(cell)
                cell.fill = INVALID_FILL
                cell.font = Font(color="9C5700")
                if col == 4:
                    cell.number_format = "#,##0.00"
            report_row += 1

    if violation_count == 0:
        ws_report.cell(row=2, column=1, value="✅ No validation violations found.")
    else:
        ws_report.cell(
            row=report_row + 1, column=1,
            value=f"Total violations: {violation_count}"
        ).font = Font(bold=True, color="C00000")

    return wb

# ─────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Loading CSV …")
    rows = load_csv(CSV_FILE)
    print(f"  {len(rows)} rows loaded")

    print("Building workbook …")
    wb = build_workbook(rows)

    wb.save(OUTPUT)
    print(f"\n✅  Saved: {OUTPUT}")
    print("\nSheets created:")
    for ws in wb.worksheets:
        state = " (hidden)" if ws.sheet_state == "hidden" else ""
        print(f"  • {ws.title}{state}")

    print("\nData Validation rules applied:")
    print("  • B (Age)       — Whole number, between 18 and 65  [STOP]")
    print("  • C (Department)— List: Sales, Marketing, Finance, Engineering, HR, IT  [STOP]")
    print("  • D (Salary)    — Decimal, between 30,000 and 100,000  [STOP]")
    print("  • D (Salary)    — Custom formula ≤2 decimal places  [WARNING]")

    print("\nConditional Formatting rules applied:")
    print("  • Colour Scale on Salary: White → Blue → Dark Navy")
    print(f"  • Data Bars   on Age:    Blue gradient bars (Excel default)")
    print(f"  • Icon Sets   on Salary: ↑ > {ICONSET_SALARY_THRESHOLD:,}  |  → = {ICONSET_SALARY_THRESHOLD:,}  |  ↓ < {ICONSET_SALARY_THRESHOLD:,}")

    # Count violations
    rows_all = [(r["Name"], r["Age"], r["Department"], r["Salary"]) for r in rows]
    for emp in NEW_EMPLOYEES:
        rows_all.append(emp)
    violations = [
        (n, a, d, s) for n, a, d, s in rows_all
        if not (18 <= a <= 65)
        or d not in VALID_DEPARTMENTS
        or not (30000 <= s <= 100000)
    ]
    print(f"\n⚠  Rows with existing violations (highlighted yellow): {len(violations)}")
    for v in violations:
        issues = []
        if not (18 <= v[1] <= 65):        issues.append(f"Age={v[1]}")
        if v[2] not in VALID_DEPARTMENTS:  issues.append(f"Dept='{v[2]}'")
        if not (30000 <= v[3] <= 100000):  issues.append(f"Salary={v[3]:,.2f}")
        print(f"     {v[0]:<25} → {', '.join(issues)}")

    print(f"\n🆕  New employee rows added (highlighted blue): {len(NEW_EMPLOYEES)}")
    for emp in NEW_EMPLOYEES:
        print(f"     {emp[0]:<25} Age={emp[1]}  Dept={emp[2]}  Salary={emp[3]:,.2f}")
